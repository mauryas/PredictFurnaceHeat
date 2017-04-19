# -*- coding: utf-8 -*-
"""
Created on Sat Nov 26 23:24:06 2016

@author: ShivamMaurya
"""

import pandas as pd
from sklearn import linear_model
from sklearn.metrics import r2_score
import matplotlib.pyplot as plt
import numpy as np
import matplotlib.backends.backend_pdf as bpdf
from sklearn.metrics import mean_squared_error
from math import sqrt
import openpyxl as opxl
#from LinearAndMultiLinearRegression import combineDay1Day2TestDay3
#from LinearAndMultiLinearRegression import combineDay2Day3TestDay1
#from LinearAndMultiLinearRegression import combineDay1Day3TestDay2


dataCompete = pd.ExcelFile('C:/MASTERS/Project/regression analysis/combinedData50000.xlsx') 
wsday1 =  dataCompete.parse('Sheet1')[0:50006]
wsday2 =  dataCompete.parse('Sheet2')[0:43168]
wsday3 =  dataCompete.parse('Sheet3')[0:43037]


#pdf file
#pdf = bpdf.PdfPages("Analysis Plots - Dynamic Linear Regression_withoutMeanDay1.pdf")
# Excel Workbook
wb = opxl.Workbook()
sheet = wb.get_active_sheet()
sheet.title = 'DynamicRegression RMS on Day3'
sheet.cell(row=1, column=1).value = "Relation"
sheet.cell(row=1, column=2).value = "RMS"
sheet.cell(row=1, column=3).value = "R-square Total Predicted"
sheet.cell(row=1, column=4).value = "R-square Static Predicted"

for x in range(1,9):
    independentDay1 = np.array(wsday1[["input_heat_energy"+str(x),"input_wall_temp"+str(x),"input_mat_temp"+str(x)]])
    dependentDay1 = np.array(wsday1["output_section_temp"+str(x)])
    independentDay2 = np.array(wsday2[["input_heat_energy"+str(x),"input_wall_temp"+str(x),"input_mat_temp"+str(x)]])
    dependentDay2 = np.array(wsday2["output_section_temp"+str(x)])
    independentDay3 = np.array(wsday3[["input_heat_energy"+str(x),"input_wall_temp"+str(x),"input_mat_temp"+str(x)]])
    dependentDay3 = np.array(wsday3["output_section_temp"+str(x)])
    relation = "output_section_temp "+str(x) + " ~ input_heat_energy"+str(x)," input_wall_temp"+str(x)," input_mat_temp"+str(x)
    independent=np.append(independentDay1,independentDay2,axis=0)
    dependent=np.append(dependentDay1,dependentDay2,axis=0)
    sheet.cell(row=x+1, column=1).value = str(relation)
    
    
    ##Initialize the model
    LinearRegressor = linear_model.LinearRegression()
    LinearRegressor.fit(independent[0:7200],dependent[0:7200])
    totalPredicted = np.array(LinearRegressor.predict(independent[0:7200]))
    staticPredict = LinearRegressor.predict(independent)
    
    
    
    learningRate = 0.0000000001
    
    for y in range(0,95):
        start = int(7200 + y*900)
        end = int(7200+(y+1)*900)
        
        gradient_0=0
        gradient_1=0
        gradient_2=0
        errorSum=0
        for z in range(start,end):
            predictOutput = LinearRegressor.predict(independent[start:start+1])
            error=dependent[start:start+1] - predictOutput
            errorSum = errorSum +learningRate*error*1
            gradient_0= gradient_0 + learningRate*error*independent[start:start+1,0:1]
            gradient_1= gradient_1 + learningRate*error*independent[start:start+1,1:2]
            gradient_2= gradient_2 + learningRate*error*independent[start:start+1,2:3]

        
        LinearRegressor.intercept_ += errorSum
        LinearRegressor.coef_[0] +=  gradient_0
        LinearRegressor.coef_[1] +=  gradient_1
        LinearRegressor.coef_[2] +=  gradient_2
        predictDelta = LinearRegressor.predict(independent[start:end])
        totalPredicted = np.append(totalPredicted,predictDelta,axis=0)
        
#    print(error,totalPredicted.shape,dependent.shape)
    print(LinearRegressor.coef_,LinearRegressor.intercept_)
    predictedDay3=LinearRegressor.predict(independentDay3)
    RMS = sqrt(mean_squared_error(dependentDay3, predictedDay3))
#    print('Linear Regression: RMS on Day3 = ', RMS)
    
    sheet.cell(row=x+1, column=2).value = RMS
    sheet.cell(row=x+1, column=3).value = r2_score(dependent[7200:92700],totalPredicted[7200:92700])
    sheet.cell(row=x+1, column=4).value = r2_score(dependent[7200:92700],staticPredict[7200:92700])
#    print(r2_score(dependent[7200:92700],totalPredicted[7200:92700]))
#    print(r2_score(dependent[7200:92700],staticPredict[7200:92700]))
    print()
    
wb.save('Analysis Plots - Dynamic Linear Regression_withoutMeanDay1Day2TestDay3.xlsx')    
#pdf.close()
    
