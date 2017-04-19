# -*- coding: utf-8 -*-
"""
Created on Sat Nov 12 23:42:42 2016

@author: ShivamMaurya
"""
import pandas
import numpy as np
#from sklearn import svm 
from sklearn import linear_model
from sklearn.neural_network import MLPRegressor
from sklearn.metrics import mean_squared_error
from math import sqrt
import matplotlib.pyplot as plt
import matplotlib.backends.backend_pdf as bpdf
import openpyxl as opxl
from sklearn.metrics import r2_score


##############################################################################
# DATA STAGING
##############################################################################
    
dataCompete = pandas.ExcelFile('Combines Data with Material temp.xlsx') 
ws =  dataCompete.parse('Sheet1')

#pdf file
pdf = bpdf.PdfPages("results\Regression Analysis Plots sgd.pdf")

# Excel Workbook
wb = opxl.Workbook()
sheet = wb.get_active_sheet()
sheet.title = 'R Square'
sheet.cell(row=1, column=1).value = 'MLP Regerssor'
sheet.cell(row=2, column=1).value = 'OLS Regerssor'

for x in range(1,17):
    outputTemp = 'outputTemp'+str(x)
    inputWall = 'inputWall'+str(x)
    inputEnergy = 'inputEnergy'+str(x)
    inputAir = 'inputAir1'
    inputMat = 'inputMat'+str(x)
    target = np.array(ws[outputTemp])
    #target = target.reshape(1,1)
    independentVariable = (np.array(ws[[inputWall,inputEnergy,inputMat,inputAir]]))
    independentVariable = independentVariable.reshape(110,4)
    
    targetTrain = target[0:74]
    independentVariableTrain = independentVariable[0:74]
    
    independentVariableTrain.reshape(74,4)
    
    targetTest = target[75:110]
    independentVariableTest = (independentVariable[75:110])
    
    independentVariableTest.reshape(35,4)
    
    # %% MLP Regression
    
    regrMLP = MLPRegressor()
    regrMLP.fit(independentVariableTrain,targetTrain)
    
    #regrOutMLP = MLPRegressor()
    #regrOutMLP.fit(independentVariableTest,targetTest)
    
    
    predictedValidateMLP = regrMLP.predict(independentVariableTest)
    # Predict from the training features only
    predictedTrainMLP = regrMLP.predict(independentVariableTrain)
    # Predict from the whole training set (training+validation features)
    predictedMLP = regrMLP.predict(independentVariable)
    
    # Compute the root mean squared (RMS) for each case above
    RMSvalid = sqrt(mean_squared_error(targetTest, predictedValidateMLP))
    RMStrain = sqrt(mean_squared_error(targetTrain, predictedTrainMLP))
    RMS = sqrt(mean_squared_error(target, predictedMLP))
    # Print to standard IO the RMS and prediction scores.
    
    print('MLP Regression: RMS on validation set '+str(x)+' = ', RMSvalid)
#    print('MLP Regression: RMS on training set '+x+' = ', RMStrain)
#    print('MLP Regression: RMS on whole set '+x+' = ', RMS)
        
    rsquareMLP= r2_score#regrMLP.score(independentVariableTest, targetTest)
    print('MLP Regression: Score on validation set '+str(x)+' = ', rsquareMLP)
#    print('MLP Regression: Score on training set '+x+' = ', regrMLP.score(independentVariableTrain, targetTrain))
#    print('MLP Regression: Score on whole set '+x+' = ', regrMLP.score(independentVariable, target))
    sheet.cell(row=1,column=x+1).value = rsquareMLP
    
    # %% Linear Regression
    regrLR = linear_model.LinearRegression()
    regrLR.fit(independentVariableTrain,targetTrain)
    
    #regrOutLR = linear_model.LinearRegression()
    #regrOutLR.fit(independentVariableTest,targetTest)
    
    predictedValidateLR = regrLR.predict(independentVariableTest)
    # Predict from the training features only
    predictedTrainLR = regrLR.predict(independentVariableTrain)
    # Predict from the whole training set (training+validation features)
    predictedLR = regrLR.predict(independentVariable)
    
    # Compute the root mean squared (RMS) for each case above
    RMSvalid = sqrt(mean_squared_error(targetTest, predictedValidateLR))
    RMStrain = sqrt(mean_squared_error(targetTrain, predictedTrainLR))
    RMS = sqrt(mean_squared_error(target, predictedLR))
    
    # Print to standard IO the RMS and prediction scores.
    
    print('Linear Regression: RMS on validation set '+str(x)+' = ', RMSvalid)
#    print('Linear Regression: RMS on training set '+x+' = ', RMStrain)
#    print('Linear Regression: RMS on whole set '+x+' = ', RMS)
    rsquareOLS = regrLR.score(independentVariableTest, targetTest)
    print('Linear Regression: Score on validation set '+str(x)+' = ', rsquareOLS)
#    print('Linear Regression: Score on training set '+x+' = ', regrLR.score(independentVariableTrain, targetTrain))
#    print('Linear Regression: Score on whole set '+x+' = ', regrLR.score(independentVariable, target))
    sheet.cell(row=2,column=x+1).value = rsquareOLS    

    plt.plot(ws['time'][75:110], targetTest, 'b-', label = 'Actual Data')
    plt.plot(ws['time'][75:110], predictedValidateMLP, 'r--', label = 'MLP Predicted')
    plt.plot(ws['time'][75:110], predictedValidateLR, 'g--', label = 'LR Predicted')
    plt.xlabel('Time')
    plt.ylabel('Output Temperature'+str(x))
    plt.savefig(pdf, format='pdf')
    plt.close()

pdf.close()
wb.save('results\MLP vs OLS sgd.xlsx')