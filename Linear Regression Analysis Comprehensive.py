# -*- coding: utf-8 -*-
"""
Created on Tue Oct 25 15:20:18 2016

@author: ShivamMaurya
"""

import pandas as pd
import statsmodels.formula.api as smf
import matplotlib.pyplot as plt
import matplotlib.backends.backend_pdf as bpdf
import openpyxl as opxl
#Reading File
dataCompete = pd.ExcelFile('Combined Data.xlsx') 
ws =  dataCompete.parse('Sheet1')

#pdf file
pdf = bpdf.PdfPages("results\Analysis Plots - Linear Regression.pdf")

# Excel Workbook
wb = opxl.Workbook()
sheet = wb.get_active_sheet()
sheet.title = 'Adjusted R Square'

#Regression Analysis
for y in range(1,7):
    for x in range(1,17):
        
    #   Different dependable variables we want to analyse for      
            if y == 1:
                relation = 'outputTemp' +str(x)+ ' ~ inputEnergy' +str(x)+ ' + inputWall' +str(x)+ ' + inputAir1'
            elif y == 2:
                relation = 'outputTemp' +str(x)+ ' ~ inputEnergy' +str(x)+ ' + inputWall' +str(x)
            elif y == 3:
                relation = 'outputTemp' +str(x)+ ' ~ inputEnergy' +str(x)
            elif y == 4:
                relation = 'outputTemp' +str(x)+ ' ~ inputAir1'
            elif y == 5:
                relation = 'outputTemp' +str(x)+ ' ~ inputWall' +str(x)
            elif y == 6:
                relation = 'outputTemp' +str(x)+ ' ~ time'
            if x == 1:
                sheet.cell(row=y, column=1).value = relation#Adj R^2
            print(relation)
            regSection = smf.ols(relation, data = ws).fit()
            
            # Excel Entry
            sheet.cell(row=y, column=x+1).value = regSection.rsquared_adj 
            
            #Creating the graph for original vs predicted values
            outputTemp = 'outputTemp' + str(x)
            plt.plot(ws['time'], ws[outputTemp], 'b-', label = 'Actual Data')
            plt.plot(ws['time'], regSection.fittedvalues, 'r--', label = 'Predicted Data')
            plt.xlabel('Time')
            graphTitle = relation   #'Zone'+str(x)+':  ('+ outputTemp + ' = ' + str(round(regSection.params[0],3)) +' + '+ str(round(regSection.params[1],3))+ '* inputEnerygy'+str(x)+ ' + '+ str(round(regSection.params[2],3)) +' * inputWallTemp'+str(x)+')'
            plt.title(graphTitle,fontsize=7)
            plt.grid(True)
            plt.ylabel('Output Temperature '+str(x))
            plt.legend(loc='best',fontsize=7)            
            
        #####Save the PDF with Graphs    
            plt.savefig(pdf, format='pdf')
            plt.close()
            del regSection
# Save and close the files            
pdf.close()
wb.save('results\Analysis Excel - Linear Regression.xlsx')