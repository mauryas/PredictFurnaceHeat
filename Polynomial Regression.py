# -*- coding: utf-8 -*-
"""
Created on Thu Dec 29 22:10:50 2016

@author: Ankurbahre
"""

import pandas as pd
import statsmodels.formula.api as smf
import matplotlib.pyplot as plt
import matplotlib.backends.backend_pdf as bpdf
import openpyxl as opxl
from sklearn.metrics import mean_squared_error
from math import sqrt

#Reading File
for z in range(1,4):
    dataCompete = pd.ExcelFile('C:\ovgu\proj data\combinedData50000.xlsx')
    if z == 1:
       trainingData = dataCompete.parse('12')[0:93174]
       testData =  dataCompete.parse('Sheet3')[0:43037]
    if z ==2:
       trainingData = dataCompete.parse('23')[0:86206]
       testData =  dataCompete.parse('Sheet1')[0:50006]
        
    if z==3:
       trainingData = dataCompete.parse('31')[0:93044]
       testData =  dataCompete.parse('Sheet2')[0:43168] 
        
                        
    #pdf file
    #==============================================================================
    pdf = bpdf.PdfPages("results\Analysis Plots - Polynomial Regression"+ str(z) +".pdf")
    # 
    #  Excel Workbook
    wb = opxl.Workbook()
    sheet1 = wb.worksheets[0] #wb.get_active_sheet()
    sheet1.title = 'Polynomial Regression'
    f = open('results\coeff'+ str(z) + '.txt', 'w')
    # #Regression Analysis
    for y in range(1,9):
        for x in range(1,9):
             
         #   Different dependable variables we want to analyse for      
                 if y == 1:
                     relation = 'output_section_temp' +str(x)+ ' ~ input_heat_energy' +str(x)+ ' + input_wall_temp' +str(x)+ ' + input_mat_temp'+ str(x)+' + input_wall_temp' +str(x)+ '**2'+' + input_heat_energy' +str(x)+'**2'+ ' + input_mat_temp'+ str(x)+'**2'+' + input_wall_temp' +str(x)+ '**3'+' + input_heat_energy' +str(x)+'**3' +' + input_mat_temp'+ str(x)+'**3' #' + inputAir1'
                 elif y == 2:
                    relation = 'output_section_temp' +str(x)+ ' ~ input_heat_energy' +str(x)+ ' + input_wall_temp' +str(x)+ ' + input_mat_temp'+ str(x)+' + input_wall_temp' +str(x)+ '**2'+' + input_heat_energy' +str(x)+'**2'+ ' + input_mat_temp'+ str(x)+'**2'
                 elif y == 3:
                    relation = 'output_section_temp' +str(x)+ ' ~ input_heat_energy' +str(x)+ ' + input_heat_energy' +str(x)+'**2'+' + input_heat_energy' +str(x)+'**3'
                 elif y == 4:
                  relation = 'output_section_temp' +str(x)+ ' ~ input_wall_temp' +str(x)+ ' + input_wall_temp' +str(x)+'**2'+ ' + input_wall_temp' +str(x)+'**3'  
                 elif y == 5:
                   relation = 'output_section_temp' +str(x)+ ' ~ input_mat_temp' +str(x)+' + input_mat_temp' +str(x)+'**2'+ ' + input_mat_temp' +str(x)+'**3' 
                 elif y == 6:
                    relation = 'output_section_temp' +str(x)+ ' ~ input_heat_energy' +str(x)+ ' + input_heat_energy' +str(x)+'**2'
                 elif y == 7:
                  relation = 'output_section_temp' +str(x)+ ' ~ input_wall_temp' +str(x)+ ' + input_wall_temp' +str(x)+'**2'
                 elif y == 8:
                   relation = 'output_section_temp' +str(x)+ ' ~ input_mat_temp' +str(x)+' + input_mat_temp' +str(x)+'**2'
                           
                 if y == 1:
                     sheet1.cell(row=y, column=1).value = 'R Squared'
                 if x== 1:   
                     sheet1.cell(row=y+1, column=1).value = relation
                     print(relation)
                 regSection = smf.ols(relation,data= trainingData).fit()
                # print regSection.summary()
                 
                 def1= str(regSection.summary())
                 f.write(def1)
     # Excel Entry
                
                 sheet1.cell(row=y+1, column=x+1).value = regSection.rsquared
                 if y == 1:
                      sheet1.cell(row=y+12, column=1).value = 'Adjusted R Squared'
                 if x == 1:
                      sheet1.cell(row=y+13, column=1).value = relation
                  
                 sheet1.cell(row=y+13, column=x+1).value = regSection.rsquared_adj
                  
                  
 
                 testSection = smf.ols(relation,data= testData).fit()
                 #print testSection.summary()
                 f.write("********************TEST SECTION***********************************")
                 abc = str(testSection.summary())
                 f.write(abc)
                 if y == 1:
                      sheet1.cell(row=y+22, column=1).value = 'Predicted R Squared'
                 if x == 1:
                      sheet1.cell(row=y+23, column=1).value = relation
                 
                 sheet1.cell(row=y+23, column=x+1).value = testSection.rsquared
                  
                 regValid = regSection.predict(testData[['input_heat_energy' +str(x),'input_wall_temp' +str(x),'input_mat_temp' +str(x)]])
                 outputTemp = 'output_section_temp' + str(x)                 
                 RMSValid = sqrt(mean_squared_error(testData[outputTemp],regValid))     
                 
                 if y == 1:
                      sheet1.cell(row=y+32, column=1).value = 'Root Mean Squared Error'
                 if x == 1:
                      sheet1.cell(row=y+33, column=1).value = relation             
                  
                 sheet1.cell(row=y+33, column=x+1).value = RMSValid
                 plt.xlabel('Time')
                 plt.ylabel('Output Temperature '+str(x))
                
                 plt.plot(testData['Time Stamp'], testData['output_section_temp'+str(x)], 'b-', label = 'Actual Data')
                 plt.plot(testData['Time Stamp'], regValid, 'r-', label = 'Predicted Data')
                 plt.legend(loc='upper left')                 
                 plt.savefig(pdf, format='pdf')
                 plt.close()
    pdf.close()
    excelName =  'results\Analysis Excel - Polynomial Regression' + str(z) + 'degree.xlsx'            
    wb.save(excelName)
    f.close()
    ##==============================================================================
